import re
import xlrd
import openpyxl
import pandas as pd
import string


wb = openpyxl.load_workbook('tweet_scraper/scraped/old.xlsx')
sheet = wb.active
cleanedTweets = []

rowI = 1 # 1 indexed AND dont want column header name ie 'text'
textCol = 1
predictCol = 2
while rowI < 274:
    rowI += 1
    textCell = sheet.cell(rowI,textCol)
    predictCell = sheet.cell(rowI, predictCol)
    text = textCell.value
    prediction = predictCell.value

    cleanedText = text.lower()
    cleanedText = re.sub(r'http.+', '', cleanedText) # remove links
    cleanedText = re.sub(r'pic\.twitter\S+', '', cleanedText) # remove links to pictures
    cleanedText = re.sub(r'@#\S*', '', cleanedText) # remove mentions and hashtags
    cleanedText = re.sub(r'[+()\"\'\\/_-]', '', cleanedText)
    cleanedText = re.sub(r'\s+', ' ', cleanedText)
    info = {
        'uncleaned': text,
        'cleaned' : cleanedText,
        'prediction' : prediction
    }
    cleanedTweets.append(info)
df = pd.DataFrame(cleanedTweets)
df = df.replace({'\n': ' '}, regex=True) # remove linebreaks in the dataframe
df = df.replace({'\t': ' '}, regex=True) # remove tabs in the dataframe
df = df.replace({'\r': ' '}, regex=True) # remove carriage return in the dataframe
df.to_excel(f"cleaned_old.xlsx", index=False)
print("**done.**")
